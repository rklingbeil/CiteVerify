"""Excel report generation using openpyxl."""

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.pipeline import VerificationReport

# Status colors (ARGB format for openpyxl)
STATUS_FILLS = {
    "verified": PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid"),
    "warning": PatternFill(start_color="FEFCE8", end_color="FEFCE8", fill_type="solid"),
    "error": PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid"),
    "unverifiable": PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid"),
}

STATUS_FONTS = {
    "verified": Font(color="16A34A", bold=True),
    "warning": Font(color="CA8A04", bold=True),
    "error": Font(color="DC2626", bold=True),
    "unverifiable": Font(color="6B7280", bold=True),
}

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def generate_excel(report: VerificationReport) -> bytes:
    """Generate an Excel workbook and return the bytes."""
    wb = Workbook()

    # ── Summary Sheet ────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"

    ws_summary["A1"] = "CiteVerify — Citation Verification Report"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.merge_cells("A1:D1")

    ws_summary["A2"] = f"Document: {report.filename}"
    ws_summary["A3"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws_summary["A2"].font = Font(color="6B7280")
    ws_summary["A3"].font = Font(color="6B7280")

    # Summary stats
    stats = [
        ("Total Citations", report.total_citations),
        ("Verified", report.verified),
        ("Warnings", report.warnings),
        ("Errors", report.errors),
        ("Unverifiable", report.unverifiable),
    ]
    for row_idx, (label, value) in enumerate(stats, start=5):
        ws_summary[f"A{row_idx}"] = label
        ws_summary[f"A{row_idx}"].font = Font(bold=True)
        ws_summary[f"B{row_idx}"] = value

    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 12

    # ── Citations Sheet ──────────────────────────────────────────
    ws = wb.create_sheet("Citations")

    headers = [
        "#",
        "Case Name",
        "Citation",
        "Status",
        "Confidence",
        "Source",
        "Court",
        "Date Filed",
        "Quoted Text",
        "Quote Accuracy",
        "Quote Diff",
        "Characterization",
        "Char. Accuracy",
        "Char. Explanation",
        "Reasoning",
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Column widths
    widths = [4, 30, 35, 12, 10, 14, 20, 12, 50, 12, 40, 50, 12, 50, 60]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Data rows
    for row_idx, cr in enumerate(report.citations, start=2):
        e = cr.extraction
        l = cr.lookup
        v = cr.verification

        values = [
            row_idx - 1,
            e.case_name,
            e.citation_text,
            v.status.upper(),
            f"{v.confidence:.0%}" if v.confidence > 0 else "",
            l.source if l.found else f"Not found ({l.status})",
            l.court or "",
            l.date_filed or "",
            e.quoted_text or "",
            v.quote_accuracy or "",
            v.quote_diff or "",
            e.characterization or "",
            v.characterization_accuracy or "",
            v.characterization_explanation or "",
            v.reasoning or "",
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = WRAP

        # Status styling
        status = v.status
        status_cell = ws.cell(row=row_idx, column=4)
        if status in STATUS_FILLS:
            status_cell.fill = STATUS_FILLS[status]
            status_cell.font = STATUS_FONTS[status]

        # Row fill based on status
        if status in STATUS_FILLS:
            for col_idx in range(1, len(headers) + 1):
                if col_idx != 4:  # Skip status column (already styled)
                    ws.cell(row=row_idx, column=col_idx).fill = STATUS_FILLS[status]

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(report.citations) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
